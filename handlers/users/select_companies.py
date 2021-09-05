from aiogram.types.callback_query import CallbackQuery
from aiogram.types.input_file import InputFile
from tiker.functions_tiker import number_conversion
from aiogram.dispatcher.storage import FSMContext
from aiogram.types.inline_keyboard import InlineKeyboardButton, InlineKeyboardMarkup
from loader import dp, db
from aiogram import types
from keyboards.default import start_button
from keyboards.inline import xls_select_company, inline_cancel_button
from middlewares.internationlization import _
from states import SelectCompanies
from re import compile
from os import remove
import openpyxl

def special_match(strg, search=compile(r'[^0-9\.\-\,\_ ]').search):
    '''Проверка ввода символов, при выборке цифр'''
    return not bool(search(strg))

dict_select_company = {
    'Капитализация ': {'selected': False, 'name': 'capitalizacion'},
    'Кол-во акций ': {'selected': False, 'name': 'count_shares'},
    'Цена акции ': {'selected': False, 'name': 'prise'},
    'Общий доход ': {'selected': False, 'name': 'profit'},
    'Чистая прибыль ': {'selected': False, 'name': 'net_profit'},
    'Активы ': {'selected': False, 'name': 'assets'},
    'Обязательства ': {'selected': False, 'name': 'liab'},
    'Акционерный капитал ': {'selected': False, 'name': 'stockholder'},
    'Дивиденды ($) ': {'selected': False, 'name': 'dividends_per_dollar'},
    'Дивиденды (%) ': {'selected': False, 'name': 'dividends_per_percent'},
    'Прибыль на акцию ': {'selected': False, 'name': 'eps'},
    'Цена\прибыль ': {'selected': False, 'name': 'pe'},
    'Входит в DOW ' : {'selected': False, 'name': 'indow'},
    'Входит в SP500 ' : {'selected': False, 'name': 'insp'},
    }

dict_select_company_ENGLISH = {
    'Capitalization ': {'selected': False, 'name': 'capitalizacion'},
    'Number of shares ': {'selected': False, 'name': 'count_shares'},
    'Prise ': {'selected': False, 'name': 'prise'},
    'Total income ': {'selected': False, 'name': 'profit'},
    'Net profit ': {'selected': False, 'name': 'net_profit'},
    'Assets ': {'selected': False, 'name': 'assets'},
    'Commitments ': {'selected': False, 'name': 'liab'},
    'Share capital ': {'selected': False, 'name': 'stockholder'},
    'Dividends ($) ': {'selected': False, 'name': 'dividends_per_dollar'},
    'Dividends (%) ': {'selected': False, 'name': 'dividends_per_percent'},
    'EPS ': {'selected': False, 'name': 'eps'},
    'P/E': {'selected': False, 'name': 'pe'},
    'In DOW ' : {'selected': False, 'name': 'indow'},
    'In SP500 ' : {'selected': False, 'name': 'insp'},
    }

dict_select_company_UKRANIAN = {
    'Капіталізація ': {'selected': False, 'name': 'capitalizacion'},
    'Кількість акцій ': {'selected': False, 'name': 'count_shares'},
    'Ціна акції ': {'selected': False, 'name': 'prise'},
    'Загальний дохід ': {'selected': False, 'name': 'profit'},
    'Чистий прибуток ': {'selected': False, 'name': 'net_profit'},
    'Активи ': {'selected': False, 'name': 'assets'},
    'Зобов`язання ': {'selected': False, 'name': 'liab'},
    'Акціонерний капітал ': {'selected': False, 'name': 'stockholder'},
    'Дивіденди ($) ': {'selected': False, 'name': 'dividends_per_dollar'},
    'Дивіденди (%) ': {'selected': False, 'name': 'dividends_per_percent'},
    'Прибуток на акцію ': {'selected': False, 'name': 'eps'},
    'Ціна\прибуток ': {'selected': False, 'name': 'pe'},
    'Входить у DOW ' : {'selected': False, 'name': 'indow'},
    'Входить у SP500 ' : {'selected': False, 'name': 'insp'},
    }

letters = ('A', 'B', 'C', 'D', 'E', 'G', 'H', 'I', 'J', 'K', 'L')
letters_coma = ('F', 'M', 'N', 'O', 'P')
tuple_of_xlsx_RUSSIAN = ("Название компании", "Тикер", "Входит в индекс", "Сектор", "Капитализация", "Количество акций", "Цена", "Общий доход", "Чистая прибыль", "Активы", "Обязательства", "Акционерный капитал", "Дивиденды ($)", "Дивиденды (%)", "Прибыль на акцию", "Цена\прибыль")
tuple_of_xlsx_ENGLISH = ("The name of the company", "Ticker", "Indexed", "Sector", "Capitalization", "Number of shares", "Price", "Total income", "Net profit", "Assets", "Commitments", "Share capital", "Dividends ($)", "Dividends (%)", "EPS", "P/E")
tuple_of_xlsx_UKRANIAN = ("Назва компанії", "Тікер", "Входить у індекс", "Сектор", "Капіталізація", "Кількість акцій", "Ціна", "Загальний дохід", "Чистий прибуток", "Активи", "Зобов'язання", "Акціонерний капітал", "Дивіденди ($)", "Дивіденди (%)", "Прибуток на акцію", "Ціна/прибуток")

def do_inline_select_company(board):
    kb = InlineKeyboardMarkup()
    for key, data in board.items():
        kb.row(
            InlineKeyboardButton(text = f"{key} {'✅' if data['selected'] else '❌'}", callback_data=data['name'])
        ),
        
    kb.row(
        InlineKeyboardButton(
            text=_('Продолжить 🆗'),
            callback_data='continue'
        ),
        InlineKeyboardButton(
            text=_('Назад \U0001F448'),
            callback_data='cancel'
        )
    )
    return kb

@dp.callback_query_handler(text='cancel', state = [SelectCompanies.enter_the_sign, SelectCompanies.chose_the_sign, SelectCompanies.enter_numbers, SelectCompanies.download_xlsx])
async def cancel_select(call: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    if data.get('msg_id'):
        await dp.bot.delete_message(chat_id=call.message.chat.id, message_id=data.get('msg_id'))
    await call.message.answer(text=_('Отмена'), reply_markup=start_button)
    await state.finish()

@dp.callback_query_handler(text='continue', state = [SelectCompanies.enter_the_sign, SelectCompanies.chose_the_sign])
async def continue_select(call: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    await dp.bot.delete_message(chat_id=call.message.chat.id, message_id=data.get('msg_id'))
    attrbut = []
    attrbut_text = ''
    for key, value in data.get('keybord').items():
        if value['selected']:
            attrbut.append(value['name'])
            attrbut_text+=key + ', '
    add_sql = ''
    if len(attrbut) != 0:
        if 'indow' in attrbut:
            attrbut_text=attrbut_text.replace('In DOW , ', '')
            attrbut_text=attrbut_text.replace('Входит в DOW , ', '')
            attrbut_text=attrbut_text.replace('Входить у DOW , ', '')
        if 'insp' in attrbut:
            attrbut_text=attrbut_text.replace('Входит в SP500 , ', '')
            attrbut_text=attrbut_text.replace('In SP500 , ', '')
            attrbut_text=attrbut_text.replace('Входить у SP500 , ', '')
        if len(attrbut)<= 2 and (('insp' in attrbut) or ('indow' in attrbut)) and ('capitalizacion' not in attrbut) and ('count_shares' not in attrbut) and ('prise' not in attrbut) and ('profit' not in attrbut) and ('net_profit' not in attrbut) and ('assets' not in attrbut) and ('liab' not in attrbut) and ('stockholder' not in attrbut) and ('dividends_per_dollar' not in attrbut) and ('dividends_per_percent' not in attrbut) and ('eps' not in attrbut) and ('pe' not in attrbut):
            if 'indow' in attrbut and 'insp' in attrbut:
                add_sql = ' indow = 1 AND insp = 1'
            elif 'insp' in attrbut:
                add_sql = ' insp = 1'
            elif 'indow' in attrbut:
                add_sql = ' indow = 1'
            ansver = db.select_tickers(attributes=attrbut, add_sql=add_sql)
            answer_user = _('Тикер / ') + _('Индекс /') + '\n'
            tickers_to_load_xlsx = []
            for elem in ansver:
                tickers_to_load_xlsx.append(elem[0])
                ans = ''
                for e in elem:
                    if type(e) == float or type(e) == int:
                        pass
                    else: 
                        ans += str(e)+ ' / '
                answer_user+=ans[:-2] + '/\n' 
            answer_user +=_('\nДля получения полной информации, загрузите EXCEL таблицу')
            if len(ansver) == 0:
                await call.message.answer(text=_('К сожалению, компаний с данынми параметрами не существует'), reply_markup=start_button)
                await state.finish()
                return 0
            if len(answer_user) > 4096:
                msg_id = []
                for x in range(0, len(answer_user), 4096):
                    if len(answer_user) - x < 4096:
                        msg = await call.message.answer(text=answer_user[x:x+4096], reply_markup=xls_select_company)
                        msg_id.append(msg.message_id)
                    else:
                        msg = await call.message.answer(text=answer_user[x:x+4096])
                        msg_id.append(msg.message_id)
            else:
                msg = await call.message.answer(text=answer_user, reply_markup=xls_select_company)
                msg_id = msg.message_id
            await state.update_data(tickers = tickers_to_load_xlsx, msg_id = msg_id)
            await SelectCompanies.download_xlsx.set()
        else:
            msg = await call.message.answer(text=_('Введите значение от и до для данных атрибутов через запятую в формате:\n"20_" - от 20 до maximum \n"_1000" - от minimum до 1000\n"20_1000" - от 20 до 1000\n\nПоследовательность:\n{attrbut_text}').format(attrbut_text=attrbut_text[:-3]), reply_markup=inline_cancel_button)
            await SelectCompanies.enter_numbers.set()
        await state.update_data(msg_id=msg.message_id, attrbut=attrbut, attrbut_text=attrbut_text)
    else:
        await call.message.answer(text=_('Вы не выбрали ни одного признака для формирования выборки'), reply_markup=start_button)
        await state.finish()
        return 0

@dp.message_handler(state = SelectCompanies.enter_numbers)
async def enter_numbers(message: types.Message, state: FSMContext):
    data = await state.get_data()
    await dp.bot.delete_message(chat_id=message.chat.id, message_id=data.get('msg_id'))
    attrbut = data.get('attrbut')
    add_sql = ''
    if special_match(message.text) == False:
        await SelectCompanies.enter_the_sign.set()
        msg = await message.answer(text=_('В тексте присутствуют некорректные символы, пожалуйста повторите ввод\nРазрешенные символы: 0 1 2 3 4 5 6 7 8 9 . , - _'), reply_markup=do_inline_select_company(data.get('keybord'))) 
        await state.update_data(msg_id = msg.message_id)
        return 0
    numbers = message.text.replace(' ', '').split(',')
    lenattrbut = len(attrbut)
    if 'indow' in attrbut:
        lenattrbut-=1
    if 'insp' in attrbut:
        lenattrbut-=1
    if lenattrbut != len(numbers):
        await SelectCompanies.enter_the_sign.set()
        msg = await message.answer(text=_('Количество выбранных для формирования списка компаний атрибутов не соответсвует количеству введенных чисел\nПожалуйста еще раз сделайте ваш выбор'), reply_markup=do_inline_select_company(data.get('keybord')))
        await state.update_data(msg_id = msg.message_id)
    else:
        for elem in attrbut: 
            if elem == 'indow' and 'insp' in attrbut:
                add_sql += ' indow = 1 AND insp = 1 AND'
                attrbut.remove(elem)
                attrbut.remove('insp')
                continue
            if elem == 'indow':
                add_sql += ' indow = 1 AND'
                attrbut.remove(elem)
                continue
            if elem == 'insp':
                add_sql += ' insp = 1 AND'
                attrbut.remove(elem)
                continue
            if bool(numbers[0]):
                FROM = numbers[0][:numbers[0].find('_')]
                TO = numbers[0][numbers[0].find('_') + 1:] 
                if FROM and TO:
                    if FROM > TO:
                        await SelectCompanies.enter_the_sign.set()
                        msg = await message.answer(text=_('Некорректные параметры, значение "от" больше значения "до"\nПожалуйста еще раз сделайте ваш выбор'), reply_markup=do_inline_select_company(data.get('keybord')))
                        await state.update_data(msg_id = msg.message_id)
                        return 0
                    add_sql += f" {elem} BETWEEN {FROM} AND {TO} AND"
                elif bool(FROM)!=False:
                    add_sql += f" {elem} > {FROM} AND"
                elif bool(TO)!=0:
                    add_sql += f" {elem} < {TO} AND"
                numbers.pop(0)
        add_sql = add_sql[:-4]
        try:
            ansver = db.select_tickers(attributes=attrbut, add_sql=add_sql)
        except:
            await SelectCompanies.enter_the_sign.set()
            msg = await message.answer(text=_('Некорректные данные\nПожалуйста еще раз сделайте ваш выбор'), reply_markup=do_inline_select_company(data.get('keybord')))
            await state.update_data(msg_id = msg.message_id)
            return 0
        answer_user = _('Тикер / ') + _('Индекс / ') + str(data.get('attrbut_text')).replace(', ', ' / ') + '\n'
        tickers_to_load_xlsx = []
        for elem in ansver:
            tickers_to_load_xlsx.append(elem[0])
            ans = ''
            for e in elem:
                if type(e) == float or type(e) == int:
                    ans+= str(number_conversion(e)) + ' / '
                else: 
                    ans += str(e)+ ' / '
            answer_user+=ans[:-2] + '/\n' 
        answer_user +=_('\nДля получения полной информации, загрузите EXCEL таблицу')
        if len(ansver) == 0:
            await message.answer(text=_('К сожалению, компаний с данынми параметрами не существует'), reply_markup=start_button)
            await state.finish()
            return 0
        if len(answer_user) > 4096:
            msg_id = []
            for x in range(0, len(answer_user), 4096):
                if len(answer_user) - x < 4096:
                    msg = await message.answer(text=answer_user[x:x+4096], reply_markup=xls_select_company)
                    msg_id.append(msg.message_id)
                else:
                    msg = await message.answer(text=answer_user[x:x+4096])
                    msg_id.append(msg.message_id)
        else:
            msg = await message.answer(text=answer_user, reply_markup=xls_select_company)
            msg_id = msg.message_id
        await state.update_data(tickers = tickers_to_load_xlsx, msg_id = msg_id)
        await SelectCompanies.download_xlsx.set()
        

@dp.message_handler(text=_('Подобрать компании 🤑'))
async def select_sign(message: types.Message, state: FSMContext):
    global language_user_select_companies
    language_user_select_companies = db.select_user_language(message.from_user.id)[0]
    await SelectCompanies.enter_the_sign.set()
    if language_user_select_companies == 'en':
        await state.update_data(keybord = dict_select_company_ENGLISH)
    elif language_user_select_companies == 'uk':
        await state.update_data(keybord = dict_select_company_UKRANIAN)
    else:
        await state.update_data(keybord = dict_select_company)
    data = await state.get_data()
    keybord = data.get('keybord')
    msg = await message.answer(_('Выберите признаки на которых будет основываться подбор'), reply_markup=do_inline_select_company(keybord))
    await state.update_data(msg_id = msg.message_id)
    
@dp.callback_query_handler(state=[SelectCompanies.enter_the_sign, SelectCompanies.chose_the_sign])
async def select_interval(call: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    keybord = data.get('keybord')
    for key, value in keybord.items():
        if value['name'] == call.data:
            if value['selected'] == False:
                value['selected'] = True
            else:
                value['selected'] = False
    await state.update_data(keybord=keybord)
    await dp.bot.delete_message(chat_id=call.message.chat.id, message_id=data.get('msg_id'))
    msg = await call.message.answer(text=call.message.text, reply_markup=do_inline_select_company(keybord))
    await state.update_data(msg_id = msg.message_id)
    await SelectCompanies.chose_the_sign.set()

@dp.callback_query_handler(text='main_table', state=SelectCompanies.download_xlsx)
async def back_to_main_table(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    msg_id = data.get('msg_id')
    if type(msg_id) == list:
        for elem in msg_id:
            await dp.bot.delete_message(chat_id=call.message.chat.id, message_id=elem)
    else:
        await dp.bot.delete_message(chat_id=call.message.chat.id, message_id=msg_id)
    await call.message.answer(_('Главная страница'), reply_markup=start_button)
    await state.finish()

@dp.callback_query_handler(text='download_table', state=SelectCompanies.download_xlsx)
async def send_xlsx(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    tickers = data.get('tickers')
    result_sql = db.download_xlsx(tickers)
    file = './xlsx/'+str(call.from_user.id)+'.xlsx'
    wb = openpyxl.Workbook()
    sheet = wb.active
    tuple_of_xlsx = ()
    if language_user_select_companies == 'en':
        tuple_of_xlsx = tuple_of_xlsx_ENGLISH
    elif language_user_select_companies == 'uk':
        tuple_of_xlsx = tuple_of_xlsx_UKRANIAN
    else: 
        tuple_of_xlsx=tuple_of_xlsx_RUSSIAN
    


    # Заголовки таблицы
    for i in range(1, len(tuple_of_xlsx) + 1):
        sheet.cell(row=1, column=i).value = tuple_of_xlsx[i-1]
    # Данные таблицы
    for index, elem in enumerate(result_sql):
        for i in range(1, len(tuple_of_xlsx) + 1):
            sheet.cell(row=index+2, column=i).value = elem[i-1]
    # Форматирование
    for letter in letters:
        for index in range(1, sheet.max_row + 1):
            sheet[letter + str(index)].style = 'Currency'
        sheet[f'{letter}1'].style = '40 % - Accent2'
    for letter in letters_coma:
        for index in range(1, sheet.max_row + 1):
            sheet[letter + str(index)].style = 'Comma'
            sheet[f'F{str(index)}'].style = 'Comma [0]'
        sheet[f'{letter}1'].style = '40 % - Accent2'
        sheet['F1'].style = '40 % - Accent2'
    # Выравнивание
    for col in sheet.columns:
        for cell in col:
            alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
            cell.alignment = alignment_obj
    dims = {}
    for row in sheet.columns:
        for cell in row:
            if cell.value:  
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value + 12
    wb.save(file)
    msg_id = data.get('msg_id')
    if type(msg_id) == list:
        for elem in msg_id:
            await dp.bot.delete_message(chat_id=call.message.chat.id, message_id=elem)
    else:
        await dp.bot.delete_message(chat_id=call.message.chat.id, message_id=msg_id)
    await call.message.answer_document(InputFile(file), reply_markup=start_button)
    await state.finish()
    remove(file)
    